import openpyxl


def read_excel_data(file_path, cell_name, column):

    # Load Workbook
    workbook_ob = openpyxl.load_workbook(file_path)
    # workbook_ob = openpyxl.load_workbook(
    #     "C:\Avik\Automation\Framework\Python\PythonSeleniumBehave\TestData\TestData.xlsx")

    # Create object of the sheet you want to work on
    sheet1_ob = workbook_ob['Sheet1']

    # Find Rows having data
    rows = sheet1_ob.max_row
    columns = sheet1_ob.max_column

    profile_index = 0
    username_index = 0
    password_index = 0
    email_index = 0
    user_row = 0

    profile = cell_name

    # Find index values
    for i in range(1, 2):
        for j in range(1, columns + 1):
            cell = sheet1_ob.cell(i, j)
            if cell.value == "Profile":
                profile_index = j
            elif cell.value == "Username":
                username_index = j
            elif cell.value == "Password":
                password_index = j
            elif cell.value == "Email":
                email_index = j

    # Locate row for required user
    for i in range(1, rows + 1):
        cell = sheet1_ob.cell(i, profile_index)
        if cell.value == profile:
            user_row = i
            break

    if column == "Username":
        username = sheet1_ob.cell(user_row, username_index)
        req_value = username.value
    elif column == "Password":
        password = sheet1_ob.cell(user_row, password_index)
        req_value = password.value
    elif column == "Email":
        email = sheet1_ob.cell(user_row, email_index)
        req_value = email.value

    return req_value


def get_username_and_password(user_role):

    # Load Workbook
    # Below path should be w.r.t. the point of execution or else execution will FAIL
    workbook_ob = openpyxl.load_workbook(".\\TestData\\TestData.xlsx")

    # Create object of the sheet you want to work on
    sheet1_ob = workbook_ob['Sheet1']

    # Find Rows having data
    rows = sheet1_ob.max_row
    columns = sheet1_ob.max_column

    profile_index = 0
    username_index = 0
    password_index = 0
    email_index = 0
    user_row = 0
    user_col = 0

    profile = user_role

    # Find index values
    for i in range(1, 2):
        for j in range(1, columns + 1):
            cell = sheet1_ob.cell(i, j)
            # print(cell.value)
            if cell.value == "Profile":
                profile_index = j
                # print("Profile Index: " + str(profile_index))
            elif cell.value == "Username":
                username_index = j
                # print("Username Index: " + str(username_index))
            elif cell.value == "Password":
                password_index = j
                # print("Password Index: " + str(password_index))
            elif cell.value == "Email":
                email_index = j
                # print("Email Index: " + str(email_index))

    # Locate row for required user
    for i in range(1, rows + 1):
        cell = sheet1_ob.cell(i, profile_index)
        if cell.value == profile:
            user_row = i
            break

    username = sheet1_ob.cell(user_row, username_index)
    password = sheet1_ob.cell(user_row, password_index)

    cred = [username.value, password.value]
    return cred


